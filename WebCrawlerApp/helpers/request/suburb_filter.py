from helpers.request.constants import * 
import random
import openpyxl

class SuburbFilter:
    def __init__(self):
        self.suburb_to_district_path = RELATION_SUBURBS_TO_DISTRICTS_PATH
        self.wb = openpyxl.load_workbook(self.suburb_to_district_path)

    def get_suburbs_list(self, district):
        ws = self.wb[district]
        return [(cell.value, ws['B'][cell.row-1].value) for cell in ws['A'] if cell.value is not None]

    @staticmethod
    def random_suburbs_count(max_suburbs=None):
        return random.randint(1, max_suburbs) if max_suburbs else random.randint(1, 4)

    @staticmethod
    def get_random_suburbs(random_number_of_suburbs, suburbs_list):
        return random.sample(suburbs_list, random_number_of_suburbs)

    @staticmethod
    def reduce_suburbs_list(suburbs_list, random_suburbs):
        return [suburb for suburb in suburbs_list if suburb not in random_suburbs]

    def generate_random_suburb_subsets(self, district):
        suburb_list = self.get_suburbs_list(district)
        iterations = 0

        while suburb_list:
            num_random_suburbs = self.random_suburbs_count(min(4, len(suburb_list)))
            random_suburbs = self.get_random_suburbs(num_random_suburbs, suburb_list)
            suburb_list = self.reduce_suburbs_list(suburb_list, random_suburbs)
            iterations += 1
            yield random_suburbs, len(suburb_list)

    def get_all_districts(self):
        return self.wb.sheetnames