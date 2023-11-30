import csv 
import pandas as pd
from openpyxl import Workbook

csv_path = "/Users/pascalschutze/Programming/University/00 Thesis/WG-Gesucht/data/input/zuordnungOrtsteileBezirke.csv"
excel_path = "/Users/pascalschutze/Programming/University/00 Thesis/WG-Gesucht/data/input/zuordnungOrtsteileBezirke.xlsx"

# Read CSV file
df = pd.read_csv(csv_path, encoding='utf-8', sep=';', header=0)

# Remove leading and trailing whitespaces from Bezirk column
df['Bezirk'] = df['Bezirk'].str.strip()

df.to_csv(csv_path, sep=';', index=False)

# Create a list of unique Bezirk values
bezirke_unique = df["Bezirk"].unique().tolist()

wb = Workbook()
wb.save(excel_path)

for bezirk in bezirke_unique: 
    wb.create_sheet(title=bezirk)

for _, row in df.iterrows():
    ortsteil = row["Ortsteil"]
    bezirk = row["Bezirk"]
    sheet = wb[bezirk]
    sheet.append([ortsteil])

if wb.sheetnames[0] == "Sheet":
    wb.remove(wb["Sheet"])

wb.sheetnames.sort()

wb.save(excel_path)