import random
import openpyxl 


# Der folgende Code soll den Filter für die Berliner Ortsteile einstellen. 
# Es sollen zufällig 1 bis 4 Ortsteile ausgewählt werden und dann aus der Liste der Berliner Ortsteile entfernt werden.

suburb_to_district_path = 'data/input/zuordnungOrtsteileBezirke.xlsx'

def get_suburbs_list(district):
    """Fetch suburbs and their corresponding numbers for the given district from Excel."""
    wb = openpyxl.load_workbook(suburb_to_district_path)
    ws = wb[district]
    return [(cell.value, ws['B'][cell.row-1].value) for cell in ws['A'] if cell.value is not None]

def random_suburbs_count(max_suburbs=None):
    """Generate a random number between 1 and 4 or up to max_suburbs."""
    return random.randint(1, max_suburbs) if max_suburbs else random.randint(1, 4)

def get_random_suburbs(random_number_of_suburbs, suburbs_list):
    """Pick random suburbs from the list."""
    return random.sample(suburbs_list, random_number_of_suburbs)

def reduce_suburbs_list(suburbs_list, random_suburbs):
    """Remove selected suburbs from the main list."""
    return [suburb for suburb in suburbs_list if suburb not in random_suburbs]

def generate_random_suburb_subsets(district): 
    """Yield random subsets of suburbs for the district, and count iterations."""
    suburb_list = get_suburbs_list(district)
    iterations = 0

    while suburb_list:
        num_random_suburbs = random_suburbs_count(min(4, len(suburb_list)))
        random_suburbs = get_random_suburbs(num_random_suburbs, suburb_list)
        suburb_list = reduce_suburbs_list(suburb_list, random_suburbs)
        iterations += 1
        yield random_suburbs, len(suburb_list)


def getalldistricts(): 
    wb = openpyxl.load_workbook(suburb_to_district_path)
    return wb.sheetnames

